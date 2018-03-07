from openpyxl import Workbook, load_workbook
import os.path
#from PDFWriter import PDFWriter
class Excel:
    def clearRange(self, ws):
        #dst_wb=openpyxl.load_workbook(dest_filename, read_only=False, keep_vba=True)
        #dst_ws=dst_wb.get_sheet_by_name(this_sheet)
        #We kee the header located in row 2, and set the rest to None
        if ws.cell('A2').value is not None: #check if already cleared content
            for row in ws.iter_rows(row_offset=1):
                for cell in row:
                    cell.value=None
        
       
    def saveData(self, data, file_path, data_sheet, fnl_ws = [], col_list = [], rng_name = []):
        if os.path.isfile(file_path):
            wb = load_workbook(filename = file_path)
            #print data_sheet
            ws = wb.get_sheet_by_name(data_sheet)
            self.clearRange(ws)
        else:
            wb = Workbook()
            ws = wb.active
            self.clearRange(ws)
	
        for i, li in enumerate(data):
            liAsList = [x for x in li]
            col_cnt = 0
            for x in liAsList:
                if col_list:
                    if col_cnt + 1 in col_list:
                        ws.cell(row = i + 2, column = col_cnt + 1).value = liAsList[col_cnt]
                        ws.cell(row = i + 2, column = col_cnt + 1).number_format = '"$"#,##0_);("$"#,##0)' #formats as currency
                        col_cnt += 1
                    else:
                        ws.cell(row = i + 2, column = col_cnt + 1).value = liAsList[col_cnt]
                        col_cnt += 1
                else:
                    ws.cell(row = i + 2, column = col_cnt + 1).value = liAsList[col_cnt]
                    col_cnt += 1
        if fnl_ws:
        	ws = wb.get_sheet_by_name(fnl_ws)
        	
        wb.save(file_path)
  
  

if __name__ == "__main__": #this is what runs at the cmd line
	print "hello world"